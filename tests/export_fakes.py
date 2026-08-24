"""A store with something in it worth putting in a sheet, and a way to read one back.

The workbooks these tests produce are read back with the library rather than inspected in memory,
because a cell's type, a hyperlink's target and a byte-order mark are all things that only exist
once a file has been written and opened again.
"""

from __future__ import annotations

import csv
from collections.abc import Iterable, Mapping, Sequence
from pathlib import Path
from typing import Any

from homescout.export import export_run
from homescout.records import ListingFields, SourceRow
from homescout.store import SourceOutcome, Store


def listing(identifier: str, **fields: Any) -> SourceRow:
    """One property, with defaults for everything a given test is not talking about."""
    values: dict[str, Any] = {
        "price": 250_000,
        "listing_status": "for_sale",
        "beds": 3,
        "baths": 2,
        "sqft": 1_800,
        "lot_sqft": 43_560,
        "year_built": 1995,
        "property_type": "single_family",
        "address_line": f"{identifier} Example Road",
        "city": "Portales",
        "state": "NM",
        "postal_code": "88130",
        "county": "Roosevelt",
        "latitude": 34.1862,
        "longitude": -103.3452,
        "listing_url": f"https://listings.example.invalid/{identifier}",
    }
    values.update(fields)
    return SourceRow(
        source="realtor",
        fields=ListingFields(**values),
        payload={"id": identifier},
        source_listing_id=identifier,
    )


class Loaded:
    """A completed run, and how to find its properties by the name the test gave them."""

    def __init__(self, run_id: str, by_source_id: Mapping[str, str]) -> None:
        self.run_id = run_id
        self.ids = dict(by_source_id)

    def __getitem__(self, source_listing_id: str) -> str:
        return self.ids[source_listing_id]


def load(store: Store, rows: Iterable[SourceRow], *, search: str = "portales") -> Loaded:
    held = list(rows)
    run = store.start_run(search)
    listing_ids = store.record_observations(run.id, "realtor", held) if held else []
    store.record_source_outcome(
        run.id, SourceOutcome(source="realtor", outcome="ok", row_count=len(held))
    )
    store.complete_run(run.id)
    return Loaded(
        run.id,
        {
            row.source_listing_id or "": listing_id
            for row, listing_id in zip(held, listing_ids, strict=False)
        },
    )


def write(store: Store, loaded: Loaded, path: Path, **kwargs: Any):
    return export_run(store, loaded.run_id, path, **kwargs)


# -- reading a workbook back ------------------------------------------------


def open_workbook(path: Path) -> Any:
    from openpyxl import load_workbook

    return load_workbook(path)


def sheet_rows(path: Path, sheet: str | None = None) -> list[list[Any]]:
    book = open_workbook(path)
    found = book[sheet] if sheet else book.active
    return [list(row) for row in found.iter_rows(values_only=True)]


def headers(path: Path) -> list[str]:
    return [str(value) for value in sheet_rows(path)[0]]


def cells(path: Path, sheet: str | None = None) -> list[list[Any]]:
    """Every cell object, so a test can ask about types and hyperlinks."""
    book = open_workbook(path)
    found = book[sheet] if sheet else book.active
    return [list(row) for row in found.iter_rows()]


def column_of(path: Path, name: str) -> list[Any]:
    """One named column's values, without its header."""
    rows = sheet_rows(path)
    index = [str(value) for value in rows[0]].index(name)
    return [row[index] for row in rows[1:]]


def delimited_rows(path: Path) -> list[list[str]]:
    with path.open(encoding="utf-8-sig", newline="") as handle:
        return [row for row in csv.reader(handle)]


def template_file(root: Path, name: str, columns: Sequence[str]) -> Path:
    directory = root / "templates"
    directory.mkdir(parents=True, exist_ok=True)
    path = directory / f"{name}.yaml"
    listed = "\n".join(f"  - {column}" for column in columns)
    path.write_text(f"columns:\n{listed}\n", encoding="utf-8")
    return path


def fingerprint(store: Store) -> dict[str, list[tuple[Any, ...]]]:
    """Every row of every table, for asking whether anything at all changed.

    Every table, not a spot check of annotations. The claim is that export writes nothing, and a
    claim about everything is checked against everything.
    """
    conn = store.connection
    tables = [
        name
        for (name,) in conn.execute(
            "SELECT name FROM sqlite_master WHERE type = 'table' AND name NOT LIKE 'sqlite_%'"
        )
    ]
    found: dict[str, list[tuple[Any, ...]]] = {}
    for table in sorted(tables):
        rows = conn.execute(f"SELECT * FROM {table}").fetchall()  # noqa: S608 - names from sqlite
        found[table] = sorted(tuple(row) for row in rows)
    return found
