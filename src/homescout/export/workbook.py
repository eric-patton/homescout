"""The spreadsheet itself: two sheets, a link on every address, and nothing that will be evaluated.

Three things here are decisions rather than mechanics.

**Every text cell has its type forced.** Not the suspicious ones. `openpyxl` records a string
beginning with `=` as a formula by default, and deciding per cell is how one gets missed.

**A property with no link gets no hyperlink at all.** A cell that is styled like a link and goes
nowhere is worse than a cell that is not styled like one, and a listing URL is source-supplied text,
so its scheme is checked the same way the digest checks one before it reaches an email.

**The file is written beside its destination and moved into place.** A disk that fills up halfway
through leaves last week's sheet where it was rather than a corrupt file with this week's name.
"""

from __future__ import annotations

import contextlib
import os
from collections.abc import Mapping, Sequence
from pathlib import Path
from typing import Any
from urllib.parse import urlparse

from ..errors import HomescoutError
from .columns import Column
from .rows import Row
from .text import for_cell, looks_like_a_formula

PROPERTIES_SHEET = "Properties"
AREAS_SHEET = "Areas"

#: Wide enough to read an address without opening the cell, narrow enough that thirty-two of them
#: still fit on a screen somebody can scroll.
DEFAULT_WIDTH = 18
WIDE_COLUMNS = {"Property", "Price History & DOM", "Town Analysis Notes", "Summary", "Notes",
                "Description", "Red Flags", "Listing URL"}
WIDE_WIDTH = 42


class ExportFailed(HomescoutError):
    """The file could not be written, said in terms of the likely cause."""


def link_target(value: object) -> str | None:
    """A listing URL, if it is one this tool is willing to make clickable.

    `http` and `https` only. A `file://` target in a spreadsheet is not a broken link, it is a
    request to whatever path it names, and a listing URL is text a listing site chose.
    """
    if not isinstance(value, str) or not value.strip():
        return None
    try:
        parsed = urlparse(value.strip())
    except ValueError:
        return None
    return value.strip() if parsed.scheme in ("http", "https") and parsed.netloc else None


def _write_cell(sheet: Any, row: int, column: int, value: object) -> Any:
    """One cell, with its type settled before anything else looks at it."""
    prepared = for_cell(value)
    cell = sheet.cell(row=row, column=column, value=prepared)
    if isinstance(prepared, str):
        # Unconditionally. See this module's own explanation.
        cell.data_type = "s"
    return cell


def build(
    columns: Sequence[Column],
    rows: Sequence[Row],
    *,
    area_notes: Sequence[Any] = (),
) -> Any:
    """The workbook, in memory. Nothing is written to disk here."""
    from openpyxl import Workbook
    from openpyxl.styles import Font

    book = Workbook()
    sheet = book.active
    sheet.title = PROPERTIES_SHEET

    bold = Font(bold=True)
    for index, column in enumerate(columns, start=1):
        cell = _write_cell(sheet, 1, index, column.name)
        cell.font = bold
        letter = sheet.cell(row=1, column=index).column_letter
        sheet.column_dimensions[letter].width = (
            WIDE_WIDTH if column.name in WIDE_COLUMNS else DEFAULT_WIDTH
        )
    sheet.freeze_panes = "A2"

    for offset, row in enumerate(rows, start=2):
        for index, column in enumerate(columns, start=1):
            cell = _write_cell(sheet, offset, index, column.value(row))
            if column.links:
                target = link_target(row.fields.listing_url)
                if target is not None and cell.value:
                    cell.hyperlink = target
                    cell.style = "Hyperlink"

    _areas(book, area_notes)
    return book


def _areas(book: Any, notes: Sequence[Any]) -> None:
    """The second sheet: what the person has written about places rather than about properties.

    Always present, even with nothing in it. A workbook whose shape depends on whether somebody has
    written a note yet is a workbook two people describe differently.
    """
    from openpyxl.styles import Font

    sheet = book.create_sheet(AREAS_SHEET)
    for index, header in enumerate(("Area", "Place", "Notes", "Updated"), start=1):
        cell = _write_cell(sheet, 1, index, header)
        cell.font = Font(bold=True)
    sheet.column_dimensions["A"].width = 12
    sheet.column_dimensions["B"].width = 28
    sheet.column_dimensions["C"].width = 80
    sheet.column_dimensions["D"].width = 22
    sheet.freeze_panes = "A2"

    ordered = sorted(notes, key=lambda n: (n.area_type, n.area_value))
    for offset, note in enumerate(ordered, start=2):
        _write_cell(sheet, offset, 1, note.area_type)
        _write_cell(sheet, offset, 2, note.area_value)
        _write_cell(sheet, offset, 3, note.notes)
        _write_cell(sheet, offset, 4, note.updated_at)


def save(book: Any, path: Path) -> None:
    """Write the workbook, atomically, or say why not in terms somebody can act on."""
    path.parent.mkdir(parents=True, exist_ok=True)
    staging = path.with_name(path.name + ".partial")
    try:
        book.save(staging)
        os.replace(staging, path)
    except PermissionError:
        _discard(staging)
        raise ExportFailed(
            f"{path} could not be written because something else has it open. "
            "That is almost always the spreadsheet itself: close it and run this again."
        ) from None
    except OSError as exc:
        _discard(staging)
        raise ExportFailed(f"{path} could not be written: {exc}") from None


def _discard(staging: Path) -> None:
    """Leave nothing behind. A failure must not leave a half-written file anywhere."""
    with contextlib.suppress(OSError):
        staging.unlink(missing_ok=True)


def empty_columns(
    columns: Sequence[Column], rows: Sequence[Row]
) -> Mapping[str, tuple[str, ...]]:
    """Which columns came out entirely empty, grouped by why.

    The reason this exists: a person opening a thirty-two column sheet with eleven blank columns has
    three different questions, and "run `homescout enrich`", "write some notes" and "no free source
    supplies this" are three different answers. Getting that wrong costs somebody an afternoon
    looking for a bug in this tool.
    """
    blank = [
        column
        for column in columns
        if all(column.value(row) in (None, "") for row in rows)
    ]
    grouped: dict[str, list[str]] = {}
    for column in blank:
        grouped.setdefault(column.origin, []).append(column.name)
    return {origin: tuple(names) for origin, names in grouped.items()}


__all__ = [
    "AREAS_SHEET",
    "PROPERTIES_SHEET",
    "ExportFailed",
    "build",
    "empty_columns",
    "link_target",
    "looks_like_a_formula",
    "save",
]
